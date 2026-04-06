"""Build a 6×9 interlinear Excel sample from 1 Nephi 1."""
import json, os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
data = json.load(open(os.path.join(BASE, '_sample_data.json'), encoding='utf-8'))

wb = Workbook()
ws = wb.active
ws.title = "1 Nephi 1"
ws.sheet_view.rightToLeft = True  # RTL sheet

# ─── Page setup for 6×9 ───
ws.page_setup.paperWidth = '6in'
ws.page_setup.paperHeight = '9in'
ws.page_setup.orientation = 'portrait'
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 0  # don't constrain vertical pages
ws.sheet_properties.pageSetUpPr.fitToPage = True
ws.page_margins.left = 0.5
ws.page_margins.right = 0.5
ws.page_margins.top = 0.4
ws.page_margins.bottom = 0.3
ws.page_margins.header = 0.2
ws.page_margins.footer = 0.15

# ─── Styles ───
heb_font = Font(name='David Libre', size=13, bold=True, color='1a2744')
gloss_font = Font(name='Crimson Pro', size=7, italic=True, color='555555')
vn_font = Font(name='Crimson Pro', size=7, bold=True, color='666666')
title_font = Font(name='David Libre', size=18, bold=True, color='1a2744')
title_eng_font = Font(name='Crimson Pro', size=10, bold=True, color='555555')
ch_font = Font(name='David Libre', size=12, bold=True, color='1a2744')
arr_font = Font(name='Crimson Pro', size=5, color='bbbbbb')
sof_font = Font(name='David Libre', size=13, bold=True, color='1a2744')

heb_align = Alignment(horizontal='center', vertical='bottom', wrap_text=False)
gloss_align = Alignment(horizontal='center', vertical='top', wrap_text=False)
title_align = Alignment(horizontal='center', vertical='center')
vn_align = Alignment(horizontal='center', vertical='bottom')

thin_border_top = Border(top=Side(style='thin', color='AAAAAA'))
thin_border_bottom = Border(bottom=Side(style='thin', color='AAAAAA'))
ch_border = Border(
    top=Side(style='thin', color='AAAAAA'),
    bottom=Side(style='thin', color='AAAAAA')
)

# ─── Layout parameters ───
# How many word-pair columns fit on the page
# 6in page - 1in margins = 5in usable = 360pt
# Each word column ~0.55in on average, plus verse number column + sof column
# Use column A for verse number, then word columns, then sof-pasuk
WORDS_PER_LINE = 8  # word-pair slots per line
VN_COL = 1          # verse number column (rightmost in RTL)
WORD_START_COL = 2  # first word column
# Arrow columns interleaved: word, arr, word, arr, ...
# Total columns needed: 1 (vn) + WORDS_PER_LINE * 2 - 1 (word+arr pairs) + 1 (sof)

# Actually simpler: no arrow columns, just word columns with arrows as part of display
# Let's use: col1=VN, col2..col(N+1)=words, col(N+2)=sof
TOTAL_COLS = VN_COL + WORDS_PER_LINE + 1  # +1 for sof-pasuk

# Set column widths
ws.column_dimensions[get_column_letter(1)].width = 3.5  # verse number
for i in range(WORDS_PER_LINE):
    ws.column_dimensions[get_column_letter(2 + i)].width = 8.0  # word columns
ws.column_dimensions[get_column_letter(2 + WORDS_PER_LINE)].width = 2.5  # sof column

row = 1

def merge_full_row(r, text, font, height=None):
    """Write text merged across all columns."""
    global row
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=TOTAL_COLS)
    cell = ws.cell(row=r, column=1, value=text)
    cell.font = font
    cell.alignment = title_align
    if height:
        ws.row_dimensions[r].height = height
    row = r + 1

# ─── Book title ───
merge_full_row(row, 'נֶפִי א׳', title_font, height=30)
merge_full_row(row, '1 Nephi', title_eng_font, height=16)
row += 1  # blank spacer row
ws.row_dimensions[row - 1].height = 6

# ─── Colophon ───
if data.get('colophon'):
    col_words = [p for p in data['colophon'] if p[0] and p[0] != '׃']
    # Render colophon as a single merged row with all words
    chunks = [col_words[i:i+WORDS_PER_LINE] for i in range(0, len(col_words), WORDS_PER_LINE)]
    for chunk in chunks:
        # Hebrew row
        for ci, pair in enumerate(chunk):
            c = ws.cell(row=row, column=2 + ci, value=pair[0])
            c.font = Font(name='David Libre', size=11, bold=True, color='1a2744')
            c.alignment = heb_align
        # Sof at end of last chunk
        if chunk == chunks[-1]:
            c = ws.cell(row=row, column=2 + len(chunk), value='׃')
            c.font = Font(name='David Libre', size=11, bold=True, color='1a2744')
            c.alignment = heb_align
        ws.row_dimensions[row].height = 16
        row += 1
        # Gloss row
        for ci, pair in enumerate(chunk):
            gloss = (pair[1] or '').replace('-', '\u2011')
            c = ws.cell(row=row, column=2 + ci, value=gloss)
            c.font = Font(name='Crimson Pro', size=6, italic=True, color='555555')
            c.alignment = gloss_align
        ws.row_dimensions[row].height = 10
        row += 1
    # spacer
    ws.row_dimensions[row].height = 4
    row += 1

# ─── Chapter heading ───
merge_full_row(row, 'פרק א', ch_font, height=20)
for c_idx in range(1, TOTAL_COLS + 1):
    cell = ws.cell(row=row - 1, column=c_idx)
    cell.border = ch_border
row += 1  # small spacer
ws.row_dimensions[row - 1].height = 4

# ─── Verses ───
for verse in data['verses']:
    words = [p for p in verse['words'] if p[0] and p[0] != '׃']
    if not words:
        continue

    # Split words into lines of WORDS_PER_LINE
    chunks = [words[i:i+WORDS_PER_LINE] for i in range(0, len(words), WORDS_PER_LINE)]

    for li, chunk in enumerate(chunks):
        # Hebrew row
        ws.row_dimensions[row].height = 18
        if li == 0:
            # Verse number in first column
            vn_cell = ws.cell(row=row, column=1, value=str(verse['num']))
            vn_cell.font = vn_font
            vn_cell.alignment = vn_align

        for ci, pair in enumerate(chunk):
            c = ws.cell(row=row, column=2 + ci, value=pair[0])
            c.font = heb_font
            c.alignment = heb_align

        # Sof-pasuk at end of last chunk
        if li == len(chunks) - 1:
            sof_col = 2 + len(chunk)
            c = ws.cell(row=row, column=sof_col, value='׃')
            c.font = sof_font
            c.alignment = heb_align

        row += 1

        # Gloss row
        ws.row_dimensions[row].height = 10
        for ci, pair in enumerate(chunk):
            gloss = (pair[1] or '').replace('-', '\u2011')
            c = ws.cell(row=row, column=2 + ci, value=gloss)
            c.font = gloss_font
            c.alignment = gloss_align

        row += 1

    # Tiny spacer between verses
    ws.row_dimensions[row].height = 3
    row += 1

# ─── Set print area so it doesn't expand ───
last_col_letter = get_column_letter(TOTAL_COLS)
ws.print_area = f'A1:{last_col_letter}{row - 1}'

# ─── Save ───
out_path = os.path.join(BASE, '_sample_1nephi1.xlsx')
wb.save(out_path)
print(f'Saved: {out_path}')
print(f'Rows: {row}')
