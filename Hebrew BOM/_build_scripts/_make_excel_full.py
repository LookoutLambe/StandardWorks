"""Build full 6x9 interlinear Excel BOM — each book on its own sheet.
   Separate Hebrew/Gloss rows with smart page breaks to keep pairs together."""
import json, os, sys
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.pagebreak import Break

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
data = json.load(open(os.path.join(BASE, '_all_data.json'), encoding='utf-8'))
front_matter = json.load(open(os.path.join(BASE, 'front_matter.json'), encoding='utf-8'))

def heb_num(n):
    if n <= 0: return str(n)
    ones = ['','א','ב','ג','ד','ה','ו','ז','ח','ט']
    tens = ['','י','כ','ל','מ','נ','ס','ע','פ','צ']
    hundreds = ['','ק','ר','ש','ת']
    if n == 15: return 'טו'
    if n == 16: return 'טז'
    r = ''
    if n >= 100: r += hundreds[n // 100]; n %= 100
    if n >= 10: r += tens[n // 10]; n %= 10
    if n > 0: r += ones[n]
    return r

wb = Workbook()

# ─── Styles ───
heb_font = Font(name='David Libre', size=13, bold=True, color='1a2744')
gloss_font = Font(name='Crimson Pro', size=7, italic=True, color='555555')
vn_font = Font(name='Crimson Pro', size=7, bold=True, color='666666')
title_font = Font(name='David Libre', size=18, bold=True, color='1a2744')
title_eng_font = Font(name='Crimson Pro', size=10, bold=True, color='555555')
ch_font = Font(name='David Libre', size=12, bold=True, color='1a2744')
col_heb_font = Font(name='David Libre', size=11, bold=True, color='1a2744')
col_gloss_font = Font(name='Crimson Pro', size=6, italic=True, color='555555')

fm_main_title_font = Font(name='David Libre', size=28, bold=True, color='1a2744')
fm_subtitle_font = Font(name='David Libre', size=14, color='333333')
fm_trans_label_font = Font(name='Crimson Pro', size=11, color='444444')
fm_trans_line_font = Font(name='David Libre', size=12, color='1a1a1a')
fm_section_title_font = Font(name='David Libre', size=15, bold=True, color='1a2744')
fm_body_font = Font(name='David Libre', size=11, color='1a1a1a')
fm_toc_font = Font(name='David Libre', size=11, color='1a1a1a')

heb_align = Alignment(horizontal='center', vertical='bottom', shrink_to_fit=True)
gloss_align = Alignment(horizontal='center', vertical='top', shrink_to_fit=True)
title_align = Alignment(horizontal='center', vertical='center')
vn_align = Alignment(horizontal='center', vertical='bottom')
fm_body_align = Alignment(horizontal='right', vertical='top', wrap_text=True)
fm_center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

ch_border = Border(
    top=Side(style='thin', color='AAAAAA'),
    bottom=Side(style='thin', color='AAAAAA')
)

WORDS_PER_LINE = 8
TOTAL_COLS = 1 + WORDS_PER_LINE

# Page capacity in pre-scale points (conservative — prevents Hebrew/Gloss splitting)
# 6x9 page, ~0.85in vertical margins, fitToWidth scale ~65% → ~576pt printed / 0.65 ≈ 886pt
# Using 780pt for safety buffer
PAGE_CAPACITY = 780


def setup_interlinear_sheet(ws):
    ws.sheet_view.rightToLeft = True
    ws.page_setup.paperWidth = '6in'
    ws.page_setup.paperHeight = '9in'
    ws.page_setup.orientation = 'portrait'
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left = 0.5
    ws.page_margins.right = 0.5
    ws.page_margins.top = 0.4
    ws.page_margins.bottom = 0.3
    ws.page_margins.header = 0.15
    ws.page_margins.footer = 0.15
    ws.print_options.horizontalCentered = True
    ws.column_dimensions[get_column_letter(1)].width = 3.0
    for i in range(WORDS_PER_LINE):
        ws.column_dimensions[get_column_letter(2 + i)].width = 12.0


def setup_frontmatter_sheet(ws):
    ws.sheet_view.rightToLeft = True
    ws.page_setup.paperWidth = '6in'
    ws.page_setup.paperHeight = '9in'
    ws.page_setup.orientation = 'portrait'
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left = 0.6
    ws.page_margins.right = 0.6
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.4
    ws.page_margins.header = 0.15
    ws.page_margins.footer = 0.15
    ws.print_options.horizontalCentered = True
    ws.column_dimensions['A'].width = 80


def merge_row(ws, r, text, font, height=None, align=None, cols=TOTAL_COLS):
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=cols)
    cell = ws.cell(row=r, column=1, value=text)
    cell.font = font
    cell.alignment = align or title_align
    if height:
        ws.row_dimensions[r].height = height


# ═══════════════════════════════════════════════════════════
# FRONT MATTER SHEETS
# ═══════════════════════════════════════════════════════════
print('Building front matter sheets...')

# Title Page
ws = wb.active
ws.title = 'Title Page'
setup_frontmatter_sheet(ws)
row = 1
ws.row_dimensions[row].height = 80; row += 1
c = ws.cell(row=row, column=1, value=front_matter[0].get('header', ''))
c.font = fm_main_title_font; c.alignment = fm_center_align
ws.row_dimensions[row].height = 50; row += 1
c = ws.cell(row=row, column=1, value='The Book of Mormon')
c.font = Font(name='Crimson Pro', size=18, bold=True, color='1a2744'); c.alignment = fm_center_align
ws.row_dimensions[row].height = 30; row += 1
ws.row_dimensions[row].height = 16; row += 1
c = ws.cell(row=row, column=1, value=front_matter[0].get('body', ''))
c.font = fm_subtitle_font; c.alignment = fm_center_align
ws.row_dimensions[row].height = 24; row += 1
c = ws.cell(row=row, column=1, value='Another Testament of Jesus Christ')
c.font = Font(name='Crimson Pro', size=11, italic=True, color='444444'); c.alignment = fm_center_align
ws.row_dimensions[row].height = 18; row += 1
ws.row_dimensions[row].height = 160; row += 1
c = ws.cell(row=row, column=1, value='Hebrew Interlinear Translation')
c.font = fm_trans_label_font; c.alignment = fm_center_align
ws.row_dimensions[row].height = 20; row += 1
for line in front_matter[1].get('full', '').split('\n'):
    line = line.strip()
    if line:
        c = ws.cell(row=row, column=1, value=line)
        c.font = fm_trans_line_font; c.alignment = fm_center_align
        ws.row_dimensions[row].height = 18; row += 1

# Sections
FM_SECTIONS = [
    (2, 'Translator Intro'), (3, 'Book of Mormon'), (4, 'Introduction'),
    (5, 'Three Witnesses'), (6, 'Eight Witnesses'), (7, 'Joseph Smith'),
    (8, 'Brief Explanation'),
]
for si, tab_name in FM_SECTIONS:
    section = front_matter[si]
    ws = wb.create_sheet(title=tab_name)
    setup_frontmatter_sheet(ws)
    row = 1
    c = ws.cell(row=row, column=1, value=section.get('header', ''))
    c.font = fm_section_title_font; c.alignment = fm_center_align
    ws.row_dimensions[row].height = 32; row += 1
    ws.row_dimensions[row].height = 8; row += 1
    for para in section.get('body', '').split('\n'):
        para = para.strip()
        if not para: continue
        c = ws.cell(row=row, column=1, value=para)
        c.font = fm_body_font; c.alignment = fm_body_align
        est_lines = max(1, len(para) // 65 + 1)
        ws.row_dimensions[row].height = est_lines * 15; row += 1

# TOC
toc_section = next((s for s in front_matter if s.get('header') == 'ראשי דברים'), None)
if toc_section:
    ws = wb.create_sheet(title='TOC')
    setup_frontmatter_sheet(ws)
    row = 1
    c = ws.cell(row=row, column=1, value='ראשי דברים')
    c.font = fm_section_title_font; c.alignment = fm_center_align
    ws.row_dimensions[row].height = 32; row += 1
    ws.row_dimensions[row].height = 10; row += 1
    for tline in toc_section.get('body', '').split('\n'):
        tline = tline.strip()
        if not tline: continue
        parts = tline.split('\t')
        book_name = parts[0].strip() if len(parts) > 0 else ''
        page_num = parts[1].strip() if len(parts) > 1 else ''
        c = ws.cell(row=row, column=1, value=f'{book_name}     {page_num}')
        c.font = fm_toc_font
        c.alignment = Alignment(horizontal='right', vertical='center')
        c.border = Border(bottom=Side(style='dotted', color='999999'))
        ws.row_dimensions[row].height = 22; row += 1


# ═══════════════════════════════════════════════════════════
# INTERLINEAR BOOK SHEETS
# ═══════════════════════════════════════════════════════════
print('Building interlinear book sheets...')
total_verses = 0

for bi, book in enumerate(data):
    book_name = book['name']
    print(f'  {book_name} ({len(book["chapters"])} chapters)...')

    ws = wb.create_sheet(title=book_name)
    setup_interlinear_sheet(ws)
    row = 1
    pg = [0]  # pg[0] = pre-scale height used on current page

    # Book title
    merge_row(ws, row, book['hebrew'], title_font, height=30)
    pg[0] += 30; row += 1
    merge_row(ws, row, book['name'], title_eng_font, height=16)
    pg[0] += 16; row += 1
    ws.row_dimensions[row].height = 6
    pg[0] += 6; row += 1

    # Colophon
    if book.get('colophon'):
        for cv in book['colophon']:
            w = cv.get('words') or cv
            if isinstance(w, list):
                filtered = [p for p in w if p[0] and p[0] != '\u05C3']
                if filtered:
                    filtered[-1] = [filtered[-1][0] + '\u05C3', filtered[-1][1]]
                    chunks = [filtered[i:i+WORDS_PER_LINE] for i in range(0, len(filtered), WORDS_PER_LINE)]
                    for chunk in chunks:
                        if pg[0] > 0 and pg[0] + 26 > PAGE_CAPACITY:
                            ws.row_breaks.append(Break(id=row - 1)); pg[0] = 0
                        ws.row_dimensions[row].height = 16
                        for ci, pair in enumerate(chunk):
                            c = ws.cell(row=row, column=2 + ci, value=pair[0])
                            c.font = col_heb_font; c.alignment = heb_align
                        pg[0] += 16; row += 1
                        ws.row_dimensions[row].height = 10
                        for ci, pair in enumerate(chunk):
                            gloss = (pair[1] or '').replace('-', '\u2011')
                            c = ws.cell(row=row, column=2 + ci, value=gloss)
                            c.font = col_gloss_font; c.alignment = gloss_align
                        pg[0] += 10; row += 1
        ws.row_dimensions[row].height = 4
        pg[0] += 4; row += 1

    for ch in book['chapters']:
        # Chapter heading — ensure heading + at least one verse line fit
        if len(book['chapters']) > 1:
            ch_need = 20 + 13 + 3 + 24 + 13  # heading + eng + spacer + one verse pair
            if pg[0] > 0 and pg[0] + ch_need > PAGE_CAPACITY:
                ws.row_breaks.append(Break(id=row - 1)); pg[0] = 0
            merge_row(ws, row, f'\u05E4\u05E8\u05E7 {heb_num(ch["num"])}', ch_font, height=20)
            for c_idx in range(1, TOTAL_COLS + 1):
                ws.cell(row=row, column=c_idx).border = Border(top=Side(style='thin', color='AAAAAA'))
            pg[0] += 20; row += 1
            merge_row(ws, row, f'Chapter {ch["num"]}', Font(name='Crimson Pro', size=8, color='555555'), height=13)
            for c_idx in range(1, TOTAL_COLS + 1):
                ws.cell(row=row, column=c_idx).border = Border(bottom=Side(style='thin', color='AAAAAA'))
            pg[0] += 13; row += 1
            ws.row_dimensions[row].height = 3
            pg[0] += 3; row += 1

        for verse in ch['verses']:
            words = [p for p in verse['words'] if p[0] and p[0] != '\u05C3']
            if not words:
                continue
            total_verses += 1

            words[-1] = [words[-1][0] + '\u05C3', words[-1][1]]
            chunks = [words[i:i+WORDS_PER_LINE] for i in range(0, len(words), WORDS_PER_LINE)]

            for li, chunk in enumerate(chunks):
                # Keep Hebrew + Gloss rows together on same page
                if pg[0] > 0 and pg[0] + 37 > PAGE_CAPACITY:
                    ws.row_breaks.append(Break(id=row - 1)); pg[0] = 0

                # Hebrew row
                ws.row_dimensions[row].height = 24
                if li == 0:
                    vn_cell = ws.cell(row=row, column=1, value=str(verse['num']))
                    vn_cell.font = vn_font; vn_cell.alignment = vn_align
                for ci, pair in enumerate(chunk):
                    c = ws.cell(row=row, column=2 + ci, value=pair[0])
                    c.font = heb_font; c.alignment = heb_align
                pg[0] += 24; row += 1

                # Gloss row
                ws.row_dimensions[row].height = 13
                for ci, pair in enumerate(chunk):
                    gloss = (pair[1] or '').replace('-', '\u2011')
                    c = ws.cell(row=row, column=2 + ci, value=gloss)
                    c.font = gloss_font; c.alignment = gloss_align
                pg[0] += 13; row += 1

            # Spacer
            ws.row_dimensions[row].height = 2
            pg[0] += 2; row += 1

    last_col = get_column_letter(TOTAL_COLS)
    ws.print_area = f'A1:{last_col}{row - 1}'

# ─── Save ───
out_path = os.path.join(os.path.expanduser('~'), 'Desktop', 'Hebrew_Interlinear_BOM.xlsx')
wb.save(out_path)
print(f'\nSaved: {out_path}')
print(f'Total verses: {total_verses}')
